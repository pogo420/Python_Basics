import openpyxl

#creating a Workbook object
wb=openpyxl.Workbook()

'''
wb.active
'''

#creating sheet
ws1 = wb.create_sheet("Mysheet",0)

'''
print(wb.sheetnames)
'''
#Addind value in row column value format
for i in range(1,101):
	for j in range(1,101):
		ws1.cell(row=i, column=j,value=i)

#Adding data in range format
ws1["A1"]="khankir chele"		

#saving a workbook
wb.save('balances.xlsx')

