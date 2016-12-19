import openpyxl
excel_name="Book1.xlsx"
sheet_name="XmlData"
file_name="XmlData.xml"
file_content=["<?xml version = \"1.0\" encoding = \"UTF-8\"?>",
				"\t<root>", 
				"\t\t<testdata>",
					"",
					"\t</testdata>", 
				"\t\t</root>",]

def convert(xmlData1,file_content1):
	dataLine="<vars"
	for i in xmlData1.keys():
		dataLine=dataLine+" "+str(i)+"="+"\""+str(xmlData1[i])+"\""
	
	dataLine=dataLine+" />"
	#print dataLine
	
	file_content1[3]="\t\t\t"+dataLine
	
	#print xmlData1
	#print file_content
	
	with open(file_name,'a') as f:
		for i in file_content1:	
			f.write(i)
			f.write("\n")
	

	


wb = openpyxl.load_workbook(excel_name)
sheet=wb.get_sheet_by_name(sheet_name)

i=1
xmlData={}

while ( sheet.cell(row=i,column=1).value != None):
	xmlData[sheet.cell(row=i,column=1).value]=sheet.cell(row=i,column=2).value
	#print sheet.cell(row=i,column=1).value,"---",sheet.cell(row=i,column=2).value
	i=i+1

convert(xmlData,file_content)

